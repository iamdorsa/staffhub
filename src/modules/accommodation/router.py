from datetime import date
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from src.core.database import get_db
from src.core.exceptions import ForbiddenError
from src.core.pagination import PaginatedResponse, PaginationParams
from src.core.permissions import CurrentUser, get_current_user, require_permission
from src.modules.accommodation import service
from src.modules.accommodation.schemas import (
    AdminReservationCreate,
    AvailabilityResponse,
    AvailabilitySetRequest,
    BannerCreate,
    BannerResponse,
    BannerUpdate,
    DiscountInfoResponse,
    OrgPlaceAccessResponse,
    OrgPlaceAccessSet,
    OrgSpecialPlanCreate,
    OrgSpecialPlanResponse,
    OrgSpecialPlanUpdate,
    PlaceCreate,
    PlaceRatingCreate,
    PlaceRatingResponse,
    PlaceRatingSummary,
    PlaceResponse,
    PlaceRoomResponse,
    PlaceRoomSet,
    PlaceUpdate,
    PricingRuleCreate,
    PricingRuleResponse,
    ReservationCreate,
    ReservationResponse,
    ReservationReviewRequest,
    RoomReservationCalendarItem,
    RoomTypeResponse,
    SpecialPlanRequestCreate,
    SpecialPlanRequestResponse,
    SpecialPlanRequestReview,
    UserPlanEligibilityResponse,
)

router = APIRouter()


# ── Room Types ────────────────────────────────────────────────────────────────

@router.get("/room-types", response_model=list[RoomTypeResponse])
def list_room_types(
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return service.list_room_types(db)


# ── Places ───────────────────────────────────────────────────────────────────

@router.get("/places", response_model=PaginatedResponse[PlaceResponse])
def list_places(
    city: Optional[str] = Query(None),
    is_active: Optional[bool] = Query(None),
    search: Optional[str] = Query(None),
    params: PaginationParams = Depends(),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return service.list_places(db, current_user, params, city=city, is_active=is_active, search=search)


@router.get("/places/{place_id}", response_model=PlaceResponse)
def get_place(
    place_id: int,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return service.get_place(db, place_id, current_user)


@router.post("/places", response_model=PlaceResponse, status_code=201)
def create_place(
    body: PlaceCreate,
    current_user: CurrentUser = Depends(require_permission("place.manage")),
    db: Session = Depends(get_db),
):
    return service.create_place(db, body)


@router.patch("/places/{place_id}", response_model=PlaceResponse)
def update_place(
    place_id: int,
    body: PlaceUpdate,
    current_user: CurrentUser = Depends(require_permission("place.manage")),
    db: Session = Depends(get_db),
):
    return service.update_place(db, place_id, body)


@router.get("/places/{place_id}/rooms", response_model=list[PlaceRoomResponse])
def list_rooms(
    place_id: int,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return service.list_rooms(db, place_id, current_user)


@router.put("/places/{place_id}/rooms", response_model=PlaceResponse)
def set_rooms(
    place_id: int,
    body: list[PlaceRoomSet],
    current_user: CurrentUser = Depends(require_permission("place.manage")),
    db: Session = Depends(get_db),
):
    return service.set_rooms(db, place_id, body)


@router.get("/places/{place_id}/availability", response_model=list[AvailabilityResponse])
def list_availability(
    place_id: int,
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    room_type_id: Optional[int] = Query(None),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return service.list_availability(db, place_id, from_date=from_date, to_date=to_date, room_type_id=room_type_id)


@router.put("/places/{place_id}/availability")
def set_availability(
    place_id: int,
    body: AvailabilitySetRequest,
    current_user: CurrentUser = Depends(require_permission("place.set_availability")),
    db: Session = Depends(get_db),
):
    return service.set_availability(db, place_id, body, admin_id=current_user.id)


@router.get("/places/{place_id}/unavailable-dates")
def list_unavailable_dates(
    place_id: int,
    from_date: date = Query(...),
    to_date: date = Query(...),
    is_vip: bool = Query(False),
    room_type_id: Optional[int] = Query(None),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    dates = service.get_unavailable_dates(db, place_id, from_date, to_date, is_vip, room_type_id=room_type_id)
    return [d.isoformat() for d in dates]


@router.get("/places/{place_id}/org-access", response_model=list[OrgPlaceAccessResponse])
def list_org_access(
    place_id: int,
    current_user: CurrentUser = Depends(require_permission("org.set_place_access")),
    db: Session = Depends(get_db),
):
    return service.list_org_access(db, place_id)


@router.put("/places/{place_id}/org-access")
def set_org_access(
    place_id: int,
    body: OrgPlaceAccessSet,
    current_user: CurrentUser = Depends(require_permission("org.set_place_access")),
    db: Session = Depends(get_db),
):
    return service.set_org_access(db, place_id, body)


# ── Reservations Calendar ───────────────────────────────────────────────────

@router.get("/places/{place_id}/reservations-calendar", response_model=list[RoomReservationCalendarItem])
def get_reservations_calendar(
    place_id: int,
    from_date: date = Query(...),
    to_date: date = Query(...),
    current_user: CurrentUser = Depends(require_permission("reservation.view_all")),
    db: Session = Depends(get_db),
):
    return service.get_reservations_calendar(db, place_id, from_date, to_date)


# ── Pricing ──────────────────────────────────────────────────────────────────

@router.post("/pricing-rules", response_model=PricingRuleResponse, status_code=201)
def create_pricing_rule(
    body: PricingRuleCreate,
    current_user: CurrentUser = Depends(require_permission("pricing.manage")),
    db: Session = Depends(get_db),
):
    return service.create_pricing_rule(db, body)


@router.get("/places/{place_id}/pricing-rules", response_model=list[PricingRuleResponse])
def list_pricing_rules(
    place_id: int,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return service.list_pricing_rules(db, place_id)


# ── Org Special Plans ────────────────────────────────────────────────────────

@router.post("/org-special-plans", response_model=OrgSpecialPlanResponse, status_code=201)
def create_org_special_plan(
    body: OrgSpecialPlanCreate,
    current_user: CurrentUser = Depends(require_permission("special_plan.manage")),
    db: Session = Depends(get_db),
):
    return service.create_org_special_plan(db, body)


@router.patch("/org-special-plans/{plan_id}", response_model=OrgSpecialPlanResponse)
def update_org_special_plan(
    plan_id: int,
    body: OrgSpecialPlanUpdate,
    current_user: CurrentUser = Depends(require_permission("special_plan.manage")),
    db: Session = Depends(get_db),
):
    return service.update_org_special_plan(db, plan_id, body)


@router.get("/orgs/{org_id}/special-plans", response_model=list[OrgSpecialPlanResponse])
def list_org_special_plans(
    org_id: int,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not current_user.is_super_admin and org_id != current_user.org_id:
        raise ForbiddenError("شما دسترسی به اطلاعات این سازمان را ندارید")
    return service.list_org_special_plans(db, org_id)


@router.get("/users/{user_id}/plan-eligibility", response_model=list[UserPlanEligibilityResponse])
def list_user_eligibility(
    user_id: int,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not current_user.is_super_admin and user_id != current_user.id:
        raise ForbiddenError("شما دسترسی به اطلاعات این کاربر را ندارید")
    return service.list_user_eligibility(db, user_id)


# ── Reservations ─────────────────────────────────────────────────────────────

@router.post("/reservations", response_model=ReservationResponse, status_code=201)
def create_reservation(
    body: ReservationCreate,
    current_user: CurrentUser = Depends(require_permission("reservation.create")),
    db: Session = Depends(get_db),
):
    return service.create_reservation(db, current_user, body)


@router.post("/reservations/admin", response_model=ReservationResponse, status_code=201)
def admin_create_reservation(
    body: AdminReservationCreate,
    current_user: CurrentUser = Depends(require_permission("reservation.approve")),
    db: Session = Depends(get_db),
):
    return service.admin_create_reservation(db, current_user, body)


@router.get("/reservations/mine", response_model=PaginatedResponse[ReservationResponse])
def list_my_reservations(
    status: Optional[str] = Query(None),
    sort_by: Optional[str] = Query(None),
    sort_dir: Optional[str] = Query(None),
    params: PaginationParams = Depends(),
    current_user: CurrentUser = Depends(require_permission("reservation.view_own")),
    db: Session = Depends(get_db),
):
    return service.list_my_reservations(db, current_user, params, status=status, sort_by=sort_by, sort_dir=sort_dir)


@router.get("/reservations", response_model=PaginatedResponse[ReservationResponse])
def list_all_reservations(
    status: Optional[str] = Query(None),
    org_id: Optional[int] = Query(None),
    place_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: Optional[str] = Query(None),
    sort_dir: Optional[str] = Query(None),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    params: PaginationParams = Depends(),
    current_user: CurrentUser = Depends(require_permission("reservation.view_all")),
    db: Session = Depends(get_db),
):
    return service.list_all_reservations(db, current_user, params, status=status, org_id=org_id, search=search, sort_by=sort_by, sort_dir=sort_dir, from_date=from_date, to_date=to_date, place_id=place_id)


@router.get("/reservations/discount-info", response_model=DiscountInfoResponse)
def get_discount_info(
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return service.get_discount_info(db, current_user.id)


@router.get("/reservations/export")
def export_reservations(
    status: Optional[str] = Query(None),
    place_id: Optional[int] = Query(None),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    search: Optional[str] = Query(None),
    current_user: CurrentUser = Depends(require_permission("reservation.view_all")),
    db: Session = Depends(get_db),
):
    rows = service.export_reservations(db, current_user, status=status, place_id=place_id, from_date=from_date, to_date=to_date, search=search)

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Reservations"
    ws.sheet_view.rightToLeft = True

    headers = ["شناسه", "کاربر", "اقامتگاه", "ورود", "خروج", "شب", "وضعیت", "VIP", "قیمت کل", "تخفیف٪", "قیمت نهایی", "طرح ویژه", "همراهان", "تاریخ ثبت"]
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_map = {"PENDING": "در انتظار", "APPROVED": "تایید شده", "REJECTED": "رد شده", "EXPIRED": "منقضی", "CANCELLED": "لغو شده"}
    for i, row in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=row["id"])
        ws.cell(row=i, column=2, value=row["user_display_name"])
        ws.cell(row=i, column=3, value=row["place_name"])
        ws.cell(row=i, column=4, value=row["check_in_date"])
        ws.cell(row=i, column=5, value=row["check_out_date"])
        ws.cell(row=i, column=6, value=row["nights"])
        ws.cell(row=i, column=7, value=status_map.get(row["status"], row["status"]))
        ws.cell(row=i, column=8, value="بله" if row["is_vip"] else "خیر")
        ws.cell(row=i, column=9, value=row["total_price"])
        ws.cell(row=i, column=10, value=row["discount_percent"])
        ws.cell(row=i, column=11, value=row["final_price"])
        ws.cell(row=i, column=12, value="بله" if row["has_special_plan"] else "خیر")
        ws.cell(row=i, column=13, value=row["guests"])
        ws.cell(row=i, column=14, value=row["created_at"])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reservations.xlsx"},
    )


@router.get("/reservations/{reservation_id}", response_model=ReservationResponse)
def get_reservation(
    reservation_id: int,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return service.get_reservation(db, reservation_id, current_user)


@router.post("/reservations/{reservation_id}/review", response_model=ReservationResponse)
def review_reservation(
    reservation_id: int,
    body: ReservationReviewRequest,
    current_user: CurrentUser = Depends(require_permission("reservation.approve")),
    db: Session = Depends(get_db),
):
    return service.review_reservation(db, reservation_id, body.action, current_user, remove_plan=body.remove_plan)


@router.post("/reservations/{reservation_id}/cancel", response_model=ReservationResponse)
def cancel_reservation(
    reservation_id: int,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return service.cancel_reservation(db, reservation_id, current_user)


# ── Ratings ─────────────────────────────────────────────────────────────────

@router.post("/place-ratings", response_model=PlaceRatingResponse, status_code=201)
def rate_place(
    body: PlaceRatingCreate,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return service.rate_place(db, current_user, body)


@router.get("/places/{place_id}/rating", response_model=PlaceRatingSummary)
def get_place_rating(
    place_id: int,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return service.get_place_rating_summary(db, place_id)


@router.get("/place-ratings/summary", response_model=list[PlaceRatingSummary])
def list_all_ratings(
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return service.list_all_place_ratings(db)


# ── Analytics ───────────────────────────────────────────────────────────────

@router.get("/analytics/places")
def place_analytics(
    current_user: CurrentUser = Depends(require_permission("place.manage")),
    db: Session = Depends(get_db),
):
    return service.get_place_analytics(db)


# ── Special Plan Requests ──────────────────────────────────────────────────

@router.post("/plan-requests", response_model=SpecialPlanRequestResponse, status_code=201)
def create_plan_request(
    body: SpecialPlanRequestCreate,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return service.create_plan_request(db, current_user, body)


@router.get("/plan-requests/mine", response_model=PaginatedResponse[SpecialPlanRequestResponse])
def list_my_plan_requests(
    params: PaginationParams = Depends(),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return service.list_my_plan_requests(db, current_user, params)


@router.get("/plan-requests", response_model=PaginatedResponse[SpecialPlanRequestResponse])
def list_plan_requests(
    status: Optional[str] = Query(None),
    params: PaginationParams = Depends(),
    current_user: CurrentUser = Depends(require_permission("special_plan.manage")),
    db: Session = Depends(get_db),
):
    return service.list_plan_requests(db, current_user, params, status=status)


@router.post("/plan-requests/{request_id}/review", response_model=SpecialPlanRequestResponse)
def review_plan_request(
    request_id: int,
    body: SpecialPlanRequestReview,
    current_user: CurrentUser = Depends(require_permission("special_plan.manage")),
    db: Session = Depends(get_db),
):
    return service.review_plan_request(db, request_id, body, current_user)


# ── Banners ────────────────────────────────────────────────────────────────

@router.get("/banners/active", response_model=Optional[BannerResponse])
def get_active_banner(
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return service.get_active_banner(db)


@router.get("/banners", response_model=list[BannerResponse])
def list_banners(
    current_user: CurrentUser = Depends(require_permission("place.manage")),
    db: Session = Depends(get_db),
):
    return service.list_banners(db)


@router.post("/banners", response_model=BannerResponse, status_code=201)
def create_banner(
    body: BannerCreate,
    current_user: CurrentUser = Depends(require_permission("place.manage")),
    db: Session = Depends(get_db),
):
    return service.create_banner(db, body, current_user.id)


@router.patch("/banners/{banner_id}", response_model=BannerResponse)
def update_banner(
    banner_id: int,
    body: BannerUpdate,
    current_user: CurrentUser = Depends(require_permission("place.manage")),
    db: Session = Depends(get_db),
):
    return service.update_banner(db, banner_id, body)


@router.delete("/banners/{banner_id}", status_code=204)
def delete_banner(
    banner_id: int,
    current_user: CurrentUser = Depends(require_permission("place.manage")),
    db: Session = Depends(get_db),
):
    service.delete_banner(db, banner_id)
